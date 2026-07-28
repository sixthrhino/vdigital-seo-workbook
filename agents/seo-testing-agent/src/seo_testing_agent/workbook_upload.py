"""Parse an uploaded .xlsx workbook (e.g. attached to a Chat message) using the
same 'On-Page' sheet layout mcp-server/tools/sheets.py reads from Google
Sheets, so Mode B works without any Sheets service-account access.

Deliberately self-contained (small, stable pure functions duplicated from
sheets.py rather than shared across services) — agent/ and mcp-server/ are
separately deployed with separate requirements.txt, and this logic is simple
enough that duplication is cheaper than wiring up a shared package.
"""

from __future__ import annotations

import datetime
import io
import re

import openpyxl

_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_month_year(text: str) -> tuple[int, int] | None:
    """Return (month, year) from various formats, or None if unparseable."""
    text = str(text).strip()
    if not text or text.lower() == "month year":
        return None
    m = re.search(r"([a-zA-Z]{3,})\s+(\d{4})", text)
    if m:
        month = _MONTH_MAP.get(m.group(1).lower()[:3])
        if month:
            return (month, int(m.group(2)))
    m = re.match(r"(\d{1,2})/\d{0,2}/?(\d{4})", text)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    m = re.match(r"(\d{4})-(\d{2})-\d{2}", text)
    if m:
        return (int(m.group(2)), int(m.group(1)))
    return None


_VOLUME_KD_RE = re.compile(r"\s*\(?[\d,]+(?:\.\d+)?[kKmM]?\s*/\s*kd\s*\d+\)?\s*$", re.I)


def _clean_keyword(kw: str) -> str:
    """Strip trailing volume/difficulty markers — parenthesized like
    '(2.5k)'/'(170)', or bare like '194000/KD 0' (some workbooks paste
    volume/KD straight from the keyword-research tool with no parens)."""
    kw = _VOLUME_KD_RE.sub("", str(kw))
    kw = re.sub(r"\s*\([^)]*\)\s*$", "", kw)
    return kw.strip()


def _is_url(text: str) -> bool:
    return bool(text) and str(text).strip().startswith("http")


def parse_on_page_rows(file_bytes: bytes, sheet: str = "On-Page", header_row: int = 4) -> list[dict]:
    """Read the raw On-Page rows from an .xlsx file's bytes, keyed by header name.

    Mirrors what gspread's ws.get_all_records(head=header_row, default_blank="")
    returns for the Google-Sheets-backed path — empty cells become "".
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in workbook. Available: {wb.sheetnames}")
    ws = wb[sheet]

    rows_iter = ws.iter_rows(min_row=header_row, values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]

    results = []
    for raw in rows_iter:
        row = {}
        for header, value in zip(headers, raw):
            if not header:
                continue
            row[header] = "" if value is None else value
        if any(v != "" for v in row.values()):
            results.append(row)
    return results


def parse_brand_guide_tab(file_bytes: bytes, sheet: str = "Brand Guide") -> str:
    """Read the uploaded workbook's Brand Guide tab into the same
    tab-separated label/value text sheets.read_brand_guide_tab produces for
    the Sheets-API path — content.py's parse_brand_guide() consumes either
    shape identically.

    Returns "" if the tab doesn't exist (not every client workbook has one).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        return ""
    ws = wb[sheet]

    lines = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if v is None else str(v) for v in row]
        while cells and cells[-1] == "":
            cells.pop()
        if cells:
            lines.append("\t".join(cells))
    return "\n".join(lines)


_CLIENT_NAME_LABELS = {"client business name", "client name", "business name"}
_WEBSITE_LABELS = {"website url", "website"}


def parse_client_details_tab(file_bytes: bytes, sheet: str = "Client Details") -> dict:
    """Read the uploaded workbook's Client Details tab for the client's
    business name and main website URL — mirrors sheets.read_client_details
    for the Sheets-API path (duplicated, not shared — see module docstring).

    Returns {"client": "", "website": ""} if the tab doesn't exist or the
    expected labels aren't found.
    """
    result = {"client": "", "website": ""}
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        return result
    ws = wb[sheet]

    for row in ws.iter_rows(values_only=True):
        if len(row) < 2 or row[0] is None or row[1] is None:
            continue
        label = str(row[0]).strip().lower().rstrip(":")
        value = str(row[1]).strip()
        if not value:
            continue
        if label in _CLIENT_NAME_LABELS:
            result["client"] = value
        elif label in _WEBSITE_LABELS:
            result["website"] = value
    return result


def _display_month(value) -> str:
    """xlsx date-typed cells come back as datetime objects, not the formatted
    strings gspread returns from Sheets — format those nicely instead of
    falling back to str(datetime(...)) ("2025-10-01 00:00:00")."""
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%B %Y")
    return str(value).strip()


def _get_col(row: dict, *names: str):
    """Return the first non-empty value among any of the given column-name
    aliases. Different client workbooks use different exact header text for
    the same field — seen in the wild: "Month Year" vs "Mo - Yr", and
    "Old Headers"/"New Headers" vs "Old H1"/"New H1". Mirrors
    tools/sheets.py::_get_col (duplicated, not shared — see module
    docstring)."""
    for name in names:
        val = row.get(name)
        if val not in (None, ""):
            return val
    return ""


def distinct_months(rows: list[dict]) -> list[str]:
    """Same distinct-month extraction as sheets.list_workbook_months, over raw rows."""
    seen: set[str] = set()
    months: list[str] = []
    for row in rows:
        val = _display_month(_get_col(row, "Mo - Yr", "Month Year"))
        if not val or val.lower() == "month year":
            continue
        parsed = _parse_month_year(val)
        if parsed and val not in seen:
            seen.add(val)
            months.append(val)
    return months


def month_rows(rows: list[dict], month_year: str) -> list[dict]:
    """Same cleaning/filtering as sheets.get_month_rows, over raw rows already in memory."""
    target = _parse_month_year(month_year)
    if target is None:
        raise ValueError(f"Could not parse month/year from: {month_year!r}")

    results = []
    for row in rows:
        mo_yr = str(_get_col(row, "Mo - Yr", "Month Year")).strip()
        if not mo_yr or mo_yr.lower() == "month year":
            continue
        if _parse_month_year(mo_yr) != target:
            continue

        url = str(row.get("Optimization / URL", "")).strip()
        if not _is_url(url):
            continue

        geo_raw = str(row.get("Target Geo", "")).strip()
        if "," in geo_raw:
            geo_city, geo_state = [p.strip() for p in geo_raw.split(",", 1)]
        else:
            geo_city, geo_state = geo_raw, ""

        visual_raw = str(row.get("Front End Visual QA", "")).strip().lower()

        results.append({
            "url": url,
            "keyword": _clean_keyword(row.get("Keyword / Volume", "")),
            "geo_city": geo_city,
            "geo_state": geo_state,
            "opt_note": str(row.get("What Is Planned / Has Been Done?", "")).strip(),
            "optimization_focus": str(row.get("Optimization Focus", "")).strip(),
            "old_title": str(row.get("Old Title Tag", "")).strip(),
            "new_title": str(row.get("New Title Tag", "")).strip(),
            "old_meta": str(row.get("Old Meta Description", "")).strip(),
            "new_meta": str(row.get("New Meta Description", "")).strip(),
            "old_h1": str(_get_col(row, "Old H1", "Old Headers")).strip(),
            "new_h1": str(_get_col(row, "New H1", "New Headers")).strip(),
            "visual_qa": visual_raw in ("true", "yes", "1"),
            "redirection": str(row.get("Redirection?", "")).strip(),
        })
    return results


def find_mentioned_month(text: str, months: list[str]) -> str | None:
    """Return the first month string from `months` that appears verbatim
    (case-insensitive) in `text`, or None if none do."""
    lowered = text.lower()
    for month in months:
        if month.lower() in lowered:
            return month
    return None
