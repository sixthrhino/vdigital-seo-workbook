"""Tests for agent/main.py — the /run endpoint, the Google Chat /chat webhook,
and the reply-truncation logic in _run_and_reply.

The ADK Runner is faked out everywhere (no real Gemini calls); only the
FastAPI routing, session wiring, and text-handling logic under test are real.
"""

import io
import json

import httpx
import openpyxl
import pytest
import respx
from starlette.testclient import TestClient

import seo_testing_agent.main as agent_main


WORKBOOK_HEADERS = [
    "Mo - Yr", "Optimization / URL", "What Is Planned / Has Been Done?",
    "Optimization Focus", "Keyword / Volume", "Target Geo", "Inspect in GSC",
    "Is It Live?", "Old Title Tag", "New Title Tag", "Old Meta Description",
    "New Meta Description", "Old H1", "New H1", "Front End Visual QA", "Redirection?",
]


def _build_test_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "On-Page"
    ws.append(["ESSENTIAL DETAILS:"])
    ws.append(["- notes"])
    ws.append(["Strategic Plan"])
    ws.append(WORKBOOK_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in WORKBOOK_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


ONE_ROW_AUGUST_XLSX = _build_test_xlsx([{
    "Mo - Yr": "August 2025",
    "Optimization / URL": "https://example.com/a",
    "What Is Planned / Has Been Done?": "Core Opts: Title Tag",
    "Keyword / Volume": "widgets",
    "Target Geo": "Phoenix, AZ",
}])


def _build_test_xlsx_with_brand_guide(rows: list[dict], brand_guide_rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "On-Page"
    ws.append(["ESSENTIAL DETAILS:"])
    ws.append(["- notes"])
    ws.append(["Strategic Plan"])
    ws.append(WORKBOOK_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in WORKBOOK_HEADERS])
    bg = wb.create_sheet("Brand Guide")
    for row in brand_guide_rows:
        bg.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


ONE_ROW_AUGUST_XLSX_WITH_BRAND_GUIDE = _build_test_xlsx_with_brand_guide(
    [{
        "Mo - Yr": "August 2025",
        "Optimization / URL": "https://example.com/a",
        "What Is Planned / Has Been Done?": "Core Opts: Title Tag",
        "Keyword / Volume": "widgets",
        "Target Geo": "Phoenix, AZ",
    }],
    [["Branding", "Acme Corp"], ["CTA", "https://example.com"]],
)


def _build_test_xlsx_with_client_details(rows: list[dict], client_details_rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "On-Page"
    ws.append(["ESSENTIAL DETAILS:"])
    ws.append(["- notes"])
    ws.append(["Strategic Plan"])
    ws.append(WORKBOOK_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in WORKBOOK_HEADERS])
    cd = wb.create_sheet("Client Details")
    for row in client_details_rows:
        cd.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


ONE_ROW_AUGUST_XLSX_WITH_CLIENT_DETAILS = _build_test_xlsx_with_client_details(
    [{
        "Mo - Yr": "August 2025",
        "Optimization / URL": "https://example.com/a",
        "What Is Planned / Has Been Done?": "Core Opts: Title Tag",
        "Keyword / Volume": "widgets",
        "Target Geo": "Phoenix, AZ",
    }],
    [["Client Business Name", "Sonoran Spine"], ["Website URL", "https://www.sonoranspine.com/"]],
)


async def _async_return(value):
    return value


# ---------------------------------------------------------------------------
# Fakes
# ---------------------------------------------------------------------------

class FakePart:
    def __init__(self, text):
        self.text = text


class FakeContent:
    def __init__(self, parts):
        self.parts = parts


class FakeEvent:
    def __init__(self, text, final=True):
        self._final = final
        self.content = FakeContent([FakePart(text)])

    def is_final_response(self):
        return self._final


def make_fake_runner(events=None, raise_exc=None, events_sequence=None):
    """events_sequence, if given, is a list of event-lists — one per
    run_async call on the same Runner instance (for testing _run_and_reply's
    retry loop). The last entry repeats for any calls beyond its length.
    Otherwise every call yields the fixed `events` list."""
    class FakeRunner:
        def __init__(self, *, agent, app_name, session_service, auto_create_session):
            self.agent = agent
            self.calls = 0

        async def run_async(self, *, user_id, session_id, new_message):
            if raise_exc:
                raise raise_exc
            if events_sequence is not None:
                idx = min(self.calls, len(events_sequence) - 1)
                self.calls += 1
                seq = events_sequence[idx]
            else:
                seq = events or []
            for e in seq:
                yield e

    return FakeRunner


class FakeCreds:
    token = "tok123"

    def refresh(self, request):
        pass


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def client(monkeypatch):
    # Force InMemorySessionService regardless of what agent/.env sets locally.
    monkeypatch.setattr(agent_main.settings, "environment", "development")
    monkeypatch.setattr(agent_main, "create_agent", lambda: object())
    with TestClient(agent_main.app) as c:
        yield c


# ---------------------------------------------------------------------------
# /health
# ---------------------------------------------------------------------------

def test_health_reports_configured_environment(client, monkeypatch):
    monkeypatch.setattr(agent_main.settings, "environment", "development")
    resp = client.get("/health")
    assert resp.status_code == 200
    assert resp.json() == {"status": "ok", "environment": "development"}


# ---------------------------------------------------------------------------
# POST /run
# ---------------------------------------------------------------------------

class TestRunEndpoint:
    def test_returns_final_agent_text(self, client, monkeypatch):
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([FakeEvent("REPORT DONE")]))
        resp = client.post("/run", json={"url": "https://example.com", "rules": ["Title Tag"]})
        assert resp.status_code == 200
        body = resp.json()
        assert body["result"] == "REPORT DONE"
        assert body["session_id"]

    def test_echoes_provided_session_id(self, client, monkeypatch):
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([FakeEvent("ok")]))
        resp = client.post("/run", json={
            "url": "https://example.com", "rules": [], "session_id": "my-session",
        })
        assert resp.json()["session_id"] == "my-session"

    def test_concatenates_multiple_final_parts(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main, "Runner",
            make_fake_runner([FakeEvent("part1"), FakeEvent("part2")]),
        )
        resp = client.post("/run", json={"url": "https://example.com", "rules": []})
        assert resp.json()["result"] == "part1part2"

    def test_ignores_non_final_events(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main, "Runner",
            make_fake_runner([FakeEvent("thinking...", final=False), FakeEvent("done")]),
        )
        resp = client.post("/run", json={"url": "https://example.com", "rules": []})
        assert resp.json()["result"] == "done"

    def test_agent_exception_returns_500(self, client, monkeypatch):
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner(raise_exc=RuntimeError("boom")))
        resp = client.post("/run", json={"url": "https://example.com", "rules": []})
        assert resp.status_code == 500
        assert "boom" in resp.json()["detail"]

    def test_invalid_url_is_rejected(self, client):
        resp = client.post("/run", json={"url": "not-a-url", "rules": []})
        assert resp.status_code == 422


class TestRunEndpointApiKey:
    def test_rejects_missing_key_when_configured(self, client, monkeypatch):
        monkeypatch.setattr(agent_main.settings, "run_api_key", "secret123")
        resp = client.post("/run", json={"url": "https://example.com", "rules": []})
        assert resp.status_code == 401

    def test_rejects_wrong_key_when_configured(self, client, monkeypatch):
        monkeypatch.setattr(agent_main.settings, "run_api_key", "secret123")
        resp = client.post(
            "/run", json={"url": "https://example.com", "rules": []},
            headers={"X-Api-Key": "wrong"},
        )
        assert resp.status_code == 401

    def test_accepts_correct_key_when_configured(self, client, monkeypatch):
        monkeypatch.setattr(agent_main.settings, "run_api_key", "secret123")
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([FakeEvent("ok")]))
        resp = client.post(
            "/run", json={"url": "https://example.com", "rules": []},
            headers={"X-Api-Key": "secret123"},
        )
        assert resp.status_code == 200

    def test_open_when_key_unconfigured(self, client, monkeypatch):
        monkeypatch.setattr(agent_main.settings, "run_api_key", "")
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([FakeEvent("ok")]))
        resp = client.post("/run", json={"url": "https://example.com", "rules": []})
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
# POST /chat
# ---------------------------------------------------------------------------

class TestChatWebhook:
    def test_added_to_space_returns_greeting(self, client):
        resp = client.post("/chat", json={"type": "ADDED_TO_SPACE"})
        assert resp.status_code == 200
        assert "ready" in resp.json()["text"].lower()

    def test_unrecognized_event_type_is_ignored(self, client):
        resp = client.post("/chat", json={"type": "REMOVED_FROM_SPACE"})
        assert resp.status_code == 200
        assert resp.json() == {}

    def test_message_without_text_is_ignored(self, client):
        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "  "},
            "space": {"name": "spaces/AAA"},
        })
        assert resp.json() == {}

    def test_message_without_space_is_ignored(self, client):
        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "hello"},
            "space": {"name": ""},
        })
        assert resp.json() == {}

    def test_message_acks_immediately_and_schedules_background_run(self, client, monkeypatch):
        calls = []

        async def fake_run_and_reply(space_name, thread_name, session_id, user_text, session_service):
            calls.append((space_name, thread_name, session_id, user_text))

        monkeypatch.setattr(agent_main, "_run_and_reply", fake_run_and_reply)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this site", "thread": {"name": "spaces/AAA/threads/T1"}},
            "space": {"name": "spaces/AAA"},
        })

        assert resp.status_code == 200
        assert "Running checks" in resp.json()["text"]
        assert calls == [("spaces/AAA", "spaces/AAA/threads/T1", "spaces-AAA-threads-T1", "review this site")]

    def test_message_without_thread_falls_back_to_space_as_session(self, client, monkeypatch):
        calls = []

        async def fake_run_and_reply(space_name, thread_name, session_id, user_text, session_service):
            calls.append((space_name, thread_name, session_id))

        monkeypatch.setattr(agent_main, "_run_and_reply", fake_run_and_reply)

        client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "hi"},
            "space": {"name": "spaces/AAA"},
        })

        assert calls == [("spaces/AAA", "spaces/AAA", "spaces-AAA")]


class TestChatWebhookDeduplication:
    """Google Chat retries a webhook delivery if it doesn't get an ack fast
    enough — confirmed live: a slow-to-process message got delivered twice
    and each delivery took a different code path, producing two conflicting
    replies in the same thread (see message.name-based dedup in handle_chat)."""

    @pytest.fixture(autouse=True)
    def clear_seen_messages(self):
        agent_main._seen_message_names.clear()
        yield
        agent_main._seen_message_names.clear()

    def test_second_delivery_of_same_message_name_is_ignored(self, client, monkeypatch):
        calls = []

        async def fake_run_and_reply(space_name, thread_name, session_id, user_text, session_service):
            calls.append(user_text)

        monkeypatch.setattr(agent_main, "_run_and_reply", fake_run_and_reply)

        payload = {
            "type": "MESSAGE",
            "message": {"name": "spaces/AAA/messages/M1", "text": "review this site"},
            "space": {"name": "spaces/AAA"},
        }

        resp1 = client.post("/chat", json=payload)
        resp2 = client.post("/chat", json=payload)

        assert resp1.json() != {}
        assert resp2.json() == {}
        assert calls == ["review this site"]

    def test_different_message_names_both_processed(self, client, monkeypatch):
        calls = []

        async def fake_run_and_reply(space_name, thread_name, session_id, user_text, session_service):
            calls.append(user_text)

        monkeypatch.setattr(agent_main, "_run_and_reply", fake_run_and_reply)

        client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"name": "spaces/AAA/messages/M1", "text": "first"},
            "space": {"name": "spaces/AAA"},
        })
        client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"name": "spaces/AAA/messages/M2", "text": "second"},
            "space": {"name": "spaces/AAA"},
        })

        assert calls == ["first", "second"]

    def test_message_without_a_name_is_never_deduped(self, client, monkeypatch):
        calls = []

        async def fake_run_and_reply(space_name, thread_name, session_id, user_text, session_service):
            calls.append(user_text)

        monkeypatch.setattr(agent_main, "_run_and_reply", fake_run_and_reply)

        payload = {
            "type": "MESSAGE",
            "message": {"text": "no name field"},
            "space": {"name": "spaces/AAA"},
        }

        client.post("/chat", json=payload)
        client.post("/chat", json=payload)

        assert calls == ["no name field", "no name field"]


class TestAlreadyProcessed:
    @pytest.fixture(autouse=True)
    def clear_seen_messages(self):
        agent_main._seen_message_names.clear()
        yield
        agent_main._seen_message_names.clear()

    def test_first_call_returns_false(self):
        assert agent_main._already_processed("spaces/A/messages/M1") is False

    def test_second_call_with_same_name_returns_true(self):
        agent_main._already_processed("spaces/A/messages/M1")
        assert agent_main._already_processed("spaces/A/messages/M1") is True

    def test_empty_name_never_counts_as_processed(self):
        assert agent_main._already_processed("") is False
        assert agent_main._already_processed("") is False

    def test_bounded_and_evicts_oldest(self, monkeypatch):
        monkeypatch.setattr(agent_main, "_SEEN_MESSAGES_MAXSIZE", 3)
        for i in range(5):
            agent_main._already_processed(f"m{i}")
        assert len(agent_main._seen_message_names) == 3
        assert "m0" not in agent_main._seen_message_names
        assert "m4" in agent_main._seen_message_names


class TestChatAuthVerification:
    def test_open_when_audience_unconfigured(self, client, monkeypatch):
        monkeypatch.setattr(agent_main.settings, "chat_audience", "")
        resp = client.post("/chat", json={"type": "ADDED_TO_SPACE"})
        assert resp.status_code == 200

    def test_rejects_missing_bearer_token_when_configured(self, client, monkeypatch):
        monkeypatch.setattr(agent_main.settings, "chat_audience", "123456789")
        resp = client.post("/chat", json={"type": "ADDED_TO_SPACE"})
        assert resp.status_code == 401

    def test_rejects_token_that_fails_verification(self, client, monkeypatch):
        monkeypatch.setattr(agent_main.settings, "chat_audience", "123456789")

        def fake_verify(token, request, audience):
            raise ValueError("invalid token")

        monkeypatch.setattr(agent_main.google_id_token, "verify_oauth2_token", fake_verify)
        resp = client.post(
            "/chat", json={"type": "ADDED_TO_SPACE"},
            headers={"Authorization": "Bearer bogus"},
        )
        assert resp.status_code == 401

    def test_rejects_token_from_wrong_issuer(self, client, monkeypatch):
        monkeypatch.setattr(agent_main.settings, "chat_audience", "123456789")

        def fake_verify(token, request, audience):
            assert audience == "123456789"
            return {"email": "someone-else@example.com", "email_verified": True}

        monkeypatch.setattr(agent_main.google_id_token, "verify_oauth2_token", fake_verify)
        resp = client.post(
            "/chat", json={"type": "ADDED_TO_SPACE"},
            headers={"Authorization": "Bearer sometoken"},
        )
        assert resp.status_code == 401

    def test_accepts_valid_chat_token(self, client, monkeypatch):
        monkeypatch.setattr(agent_main.settings, "chat_audience", "123456789")

        def fake_verify(token, request, audience):
            return {"email": "chat@system.gserviceaccount.com", "email_verified": True}

        monkeypatch.setattr(agent_main.google_id_token, "verify_oauth2_token", fake_verify)
        resp = client.post(
            "/chat", json={"type": "ADDED_TO_SPACE"},
            headers={"Authorization": "Bearer realtoken"},
        )
        assert resp.status_code == 200
        assert "ready" in resp.json()["text"].lower()


ATTACHMENT = {
    "contentName": "workbook.xlsx",
    "attachmentDataRef": {"resourceName": "spaces/AAA/messages/BBB/attachments/CCC"},
}


class TestChatWebhookWorkbookUpload:
    @pytest.fixture(autouse=True)
    def clear_pending_uploads(self):
        agent_main._pending_uploads.clear()
        yield
        agent_main._pending_uploads.clear()

    def test_attachment_with_brand_guide_tab_passes_it_through(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main, "_download_chat_attachment",
            lambda resource_name: _async_return(ONE_ROW_AUGUST_XLSX_WITH_BRAND_GUIDE),
        )
        parsed_guide = {"raw": "Branding\tAcme Corp\nCTA\thttps://example.com", "branding": ["Acme Corp"]}
        monkeypatch.setattr(
            agent_main.check_orchestrator, "parse_brand_guide_text",
            lambda text, mcp_url, auth_headers: _async_return(parsed_guide),
        )
        calls = []

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            calls.append(brand_guide)

        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "please review August 2025", "attachment": [ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert resp.status_code == 200
        assert calls == [parsed_guide]

    def test_attachment_without_brand_guide_tab_passes_empty_dict(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main, "_download_chat_attachment",
            lambda resource_name: _async_return(ONE_ROW_AUGUST_XLSX),
        )
        calls = []

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            calls.append(brand_guide)

        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "please review August 2025", "attachment": [ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert resp.status_code == 200
        assert calls == [{}]

    def test_attachment_with_client_details_tab_uses_business_name(self, client, monkeypatch):
        # Client Details' business name ("Sonoran Spine") wins over the
        # filename heuristic, which would otherwise turn ATTACHMENT's
        # contentName ("workbook.xlsx") into the useless label "workbook".
        monkeypatch.setattr(
            agent_main, "_download_chat_attachment",
            lambda resource_name: _async_return(ONE_ROW_AUGUST_XLSX_WITH_CLIENT_DETAILS),
        )
        calls = []

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            calls.append(client_name)

        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "please review August 2025", "attachment": [ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert resp.status_code == 200
        assert calls == ["Sonoran Spine"]

    def test_pending_upload_caches_brand_guide_for_followup(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main, "_download_chat_attachment",
            lambda resource_name: _async_return(ONE_ROW_AUGUST_XLSX_WITH_BRAND_GUIDE),
        )
        parsed_guide = {"raw": "Branding\tAcme Corp\nCTA\thttps://example.com", "branding": ["Acme Corp"]}
        monkeypatch.setattr(
            agent_main.check_orchestrator, "parse_brand_guide_text",
            lambda text, mcp_url, auth_headers: _async_return(parsed_guide),
        )
        client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this please", "attachment": [ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })
        raw_rows, brand_guide, client_name = agent_main._pending_uploads["spaces/AAA"]
        assert brand_guide == parsed_guide

        calls = []

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            calls.append(brand_guide)

        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "August 2025"},
            "space": {"name": "spaces/AAA"},
        })

        assert resp.status_code == 200
        assert calls == [parsed_guide]

    def test_attachment_with_month_in_text_runs_immediately(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main, "_download_chat_attachment",
            lambda resource_name: _async_return(ONE_ROW_AUGUST_XLSX),
        )
        calls = []

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            calls.append((month_year, rows))

        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "please review August 2025", "attachment": [ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert resp.status_code == 200
        assert "August 2025" in resp.json()["text"]
        assert "1 rows" in resp.json()["text"]
        assert len(calls) == 1
        month_year, rows = calls[0]
        assert month_year == "August 2025"
        assert rows[0]["url"] == "https://example.com/a"
        assert agent_main._pending_uploads == {}

    def test_attachment_without_month_asks_and_caches(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main, "_download_chat_attachment",
            lambda resource_name: _async_return(ONE_ROW_AUGUST_XLSX),
        )
        calls = []
        monkeypatch.setattr(
            agent_main, "_run_workbook_batch",
            lambda *a: calls.append(a) or _async_return(None),
        )

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this please", "attachment": [ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert "which month" in resp.json()["text"].lower()
        assert "August 2025" in resp.json()["text"]
        assert calls == []
        assert "spaces/AAA" in agent_main._pending_uploads

    def test_followup_text_resolves_pending_upload(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main, "_download_chat_attachment",
            lambda resource_name: _async_return(ONE_ROW_AUGUST_XLSX),
        )
        client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this please", "attachment": [ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })
        assert "spaces/AAA" in agent_main._pending_uploads

        calls = []

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            calls.append((month_year, rows))

        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "August 2025"},
            "space": {"name": "spaces/AAA"},
        })

        assert "August 2025" in resp.json()["text"]
        assert len(calls) == 1
        month_year, rows = calls[0]
        assert rows[0]["url"] == "https://example.com/a"
        assert agent_main._pending_uploads == {}

    def test_followup_in_a_different_thread_still_resolves(self, client, monkeypatch):
        # Same real bug as the shared-sheet equivalent test — a plain
        # top-level follow-up reply gets its own new thread ID from Chat.
        monkeypatch.setattr(
            agent_main, "_download_chat_attachment",
            lambda resource_name: _async_return(ONE_ROW_AUGUST_XLSX),
        )
        client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this please", "attachment": [ATTACHMENT],
                        "thread": {"name": "spaces/AAA/threads/T1"}},
            "space": {"name": "spaces/AAA"},
        })
        assert "spaces/AAA" in agent_main._pending_uploads

        run_and_reply_calls = []
        batch_calls = []

        async def fake_run_and_reply(space_name, thread_name, session_id, user_text, session_service):
            run_and_reply_calls.append(user_text)

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            batch_calls.append((month_year, rows))

        monkeypatch.setattr(agent_main, "_run_and_reply", fake_run_and_reply)
        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "August 2025", "thread": {"name": "spaces/AAA/threads/T2"}},
            "space": {"name": "spaces/AAA"},
        })

        assert "August 2025" in resp.json()["text"]
        assert len(batch_calls) == 1
        assert run_and_reply_calls == []
        assert agent_main._pending_uploads == {}

    def test_attachment_without_resource_name_returns_friendly_error(self, client, monkeypatch):
        calls = []
        monkeypatch.setattr(
            agent_main, "_run_workbook_batch",
            lambda *a: calls.append(a) or _async_return(None),
        )

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this", "attachment": [{"contentName": "linked.xlsx"}]},
            "space": {"name": "spaces/AAA"},
        })

        assert "can't read" in resp.json()["text"].lower()
        assert calls == []

    def test_attachment_parse_failure_returns_friendly_error(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main, "_download_chat_attachment",
            lambda resource_name: _async_return(b"not a real xlsx file"),
        )
        calls = []
        monkeypatch.setattr(
            agent_main, "_run_workbook_batch",
            lambda *a: calls.append(a) or _async_return(None),
        )

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this", "attachment": [ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert "couldn't read that workbook" in resp.json()["text"].lower()
        assert calls == []
        assert agent_main._pending_uploads == {}


DRIVE_ATTACHMENT = {
    # No contentName — matches the real payload observed live: Drive-link
    # attachments don't reliably carry a display name the way an uploaded
    # file does, which is exactly why get_workbook_title exists.
    "driveDataRef": {"driveFileId": "1L7bJ4sSU9ueao62cQu1zue1SzlBVwTwB4IqL8dVGjPM"},
}


class TestChatWebhookSharedSheet:
    """A Google Sheets link pasted/shared in Chat — represented as a Drive
    file reference (driveDataRef), not an uploaded file (attachmentDataRef).
    Mirrors TestChatWebhookWorkbookUpload but reads via the Sheets API
    (check_orchestrator.list_workbook_months/get_workbook_month_rows)
    instead of downloading and parsing .xlsx bytes."""

    @pytest.fixture(autouse=True)
    def clear_pending_sheets(self):
        agent_main._pending_sheets.clear()
        yield
        agent_main._pending_sheets.clear()

    def test_shared_sheet_with_month_in_text_runs_immediately(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main.check_orchestrator, "list_workbook_months",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(["August 2025"]),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_title",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return("IEC Rocky Mountain Workbook"),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_month_rows",
            lambda spreadsheet_id, month_year, mcp_url, auth_headers: _async_return(
                [{"url": "https://example.com/a"}]
            ),
        )
        parsed_guide = {"raw": "Branding\tAcme Corp", "branding": ["Acme Corp"]}
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_brand_guide",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(parsed_guide),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_client_details",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(
                {"client": "IEC Rocky Mountain", "website": "https://iecrm.org"}
            ),
        )
        calls = []

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            calls.append((month_year, rows, brand_guide, client_name))

        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "please review August 2025", "attachment": [DRIVE_ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert resp.status_code == 200
        assert "August 2025" in resp.json()["text"]
        assert "1 rows" in resp.json()["text"]
        assert len(calls) == 1
        month_year, rows, brand_guide, client_name = calls[0]
        assert month_year == "August 2025"
        assert rows[0]["url"] == "https://example.com/a"
        assert brand_guide == parsed_guide
        # Client Details tab's business name wins over the workbook-title heuristic.
        assert client_name == "IEC Rocky Mountain"
        assert agent_main._pending_sheets == {}

    def test_shared_sheet_brand_guide_fetch_failure_defaults_to_empty(self, client, monkeypatch):
        # A Brand Guide fetch failure (e.g. no such tab, transient MCP
        # error) shouldn't block the whole batch run — it just runs without
        # brand-guide checks, same as a workbook with no Brand Guide tab.
        monkeypatch.setattr(
            agent_main.check_orchestrator, "list_workbook_months",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(["August 2025"]),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_title",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return("IEC Rocky Mountain Workbook"),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_month_rows",
            lambda spreadsheet_id, month_year, mcp_url, auth_headers: _async_return(
                [{"url": "https://example.com/a"}]
            ),
        )

        async def boom(spreadsheet_id, mcp_url, auth_headers):
            raise RuntimeError("no Brand Guide tab")

        monkeypatch.setattr(agent_main.check_orchestrator, "get_workbook_brand_guide", boom)
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_client_details",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return({}),
        )

        calls = []

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            calls.append(brand_guide)

        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "please review August 2025", "attachment": [DRIVE_ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert resp.status_code == 200
        assert calls == [{}]

    def test_shared_sheet_without_month_asks_and_caches(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main.check_orchestrator, "list_workbook_months",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(["August 2025"]),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_title",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return("IEC Rocky Mountain Workbook"),
        )
        calls = []
        monkeypatch.setattr(
            agent_main, "_run_workbook_batch",
            lambda *a: calls.append(a) or _async_return(None),
        )

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this please", "attachment": [DRIVE_ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert "which month" in resp.json()["text"].lower()
        assert "August 2025" in resp.json()["text"]
        assert calls == []
        assert agent_main._pending_sheets["spaces/AAA"] == "1L7bJ4sSU9ueao62cQu1zue1SzlBVwTwB4IqL8dVGjPM"

    def test_uses_fetched_title_instead_of_missing_content_name(self, client, monkeypatch):
        # The real bug: DRIVE_ATTACHMENT has no contentName, so before
        # get_workbook_title existed this said 'Got "that file"'.
        monkeypatch.setattr(
            agent_main.check_orchestrator, "list_workbook_months",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(["August 2025"]),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_title",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return("IEC Rocky Mountain (main) | Organic SEO Workbook"),
        )

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this please", "attachment": [DRIVE_ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert 'Got "IEC Rocky Mountain (main) | Organic SEO Workbook"' in resp.json()["text"]
        assert "that file" not in resp.json()["text"]

    def test_falls_back_to_that_file_when_title_fetch_fails(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main.check_orchestrator, "list_workbook_months",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(["August 2025"]),
        )

        async def boom(spreadsheet_id, mcp_url, auth_headers):
            raise RuntimeError("permission denied")

        monkeypatch.setattr(agent_main.check_orchestrator, "get_workbook_title", boom)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this please", "attachment": [DRIVE_ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert 'Got "that file"' in resp.json()["text"]

    def test_followup_text_resolves_pending_sheet(self, client, monkeypatch):
        monkeypatch.setattr(
            agent_main.check_orchestrator, "list_workbook_months",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(["August 2025"]),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_title",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return("IEC Rocky Mountain Workbook"),
        )
        client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this please", "attachment": [DRIVE_ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })
        assert "spaces/AAA" in agent_main._pending_sheets

        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_month_rows",
            lambda spreadsheet_id, month_year, mcp_url, auth_headers: _async_return(
                [{"url": "https://example.com/a"}]
            ),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_brand_guide",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return({}),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_client_details",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return({}),
        )
        calls = []

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            calls.append((month_year, rows))

        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "August 2025"},
            "space": {"name": "spaces/AAA"},
        })

        assert "August 2025" in resp.json()["text"]
        assert len(calls) == 1
        assert agent_main._pending_sheets == {}

    def test_followup_in_a_different_thread_still_resolves(self, client, monkeypatch):
        # The actual bug, confirmed live: a plain top-level reply (not an
        # explicit "reply in thread") gets a brand-new thread ID from Chat,
        # even though it's clearly a follow-up in the same conversation.
        # The pending-sheet lookup has to be space-scoped, not thread-scoped,
        # or this exact sequence falls through to the generic LLM path with
        # no context.
        monkeypatch.setattr(
            agent_main.check_orchestrator, "list_workbook_months",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(["August 2025"]),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_title",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return("IEC Rocky Mountain Workbook"),
        )
        client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this please", "attachment": [DRIVE_ATTACHMENT],
                        "thread": {"name": "spaces/AAA/threads/T1"}},
            "space": {"name": "spaces/AAA"},
        })
        assert "spaces/AAA" in agent_main._pending_sheets

        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_month_rows",
            lambda spreadsheet_id, month_year, mcp_url, auth_headers: _async_return(
                [{"url": "https://example.com/a"}]
            ),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_brand_guide",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return({}),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_client_details",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return({}),
        )
        run_and_reply_calls = []
        batch_calls = []

        async def fake_run_and_reply(space_name, thread_name, session_id, user_text, session_service):
            run_and_reply_calls.append(user_text)

        async def fake_run_workbook_batch(space_name, thread_name, session_id, month_year, rows, session_service, brand_guide=None, client_name=""):
            batch_calls.append((month_year, rows))

        monkeypatch.setattr(agent_main, "_run_and_reply", fake_run_and_reply)
        monkeypatch.setattr(agent_main, "_run_workbook_batch", fake_run_workbook_batch)

        # A brand-new thread — T2, not T1 — just like the real bug.
        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "August 2025", "thread": {"name": "spaces/AAA/threads/T2"}},
            "space": {"name": "spaces/AAA"},
        })

        assert "August 2025" in resp.json()["text"]
        assert batch_calls == [("August 2025", [{"url": "https://example.com/a"}])]
        assert run_and_reply_calls == []  # must NOT have fallen through to the generic LLM path
        assert agent_main._pending_sheets == {}

    def test_list_months_failure_returns_friendly_error(self, client, monkeypatch):
        async def boom(spreadsheet_id, mcp_url, auth_headers):
            raise RuntimeError("permission denied")

        monkeypatch.setattr(agent_main.check_orchestrator, "list_workbook_months", boom)
        calls = []
        monkeypatch.setattr(
            agent_main, "_run_workbook_batch",
            lambda *a: calls.append(a) or _async_return(None),
        )

        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this", "attachment": [DRIVE_ATTACHMENT]},
            "space": {"name": "spaces/AAA"},
        })

        assert "couldn't read" in resp.json()["text"].lower()
        assert calls == []

    def test_attachment_with_neither_ref_falls_back_to_upload_error(self, client, monkeypatch):
        resp = client.post("/chat", json={
            "type": "MESSAGE",
            "message": {"text": "review this", "attachment": [{"contentName": "mystery"}]},
            "space": {"name": "spaces/AAA"},
        })

        assert "can't read" in resp.json()["text"].lower()


# ---------------------------------------------------------------------------
# _run_workbook_batch
# ---------------------------------------------------------------------------

class TestClientNameFromTitle:
    def test_strips_pipe_suffix(self):
        assert agent_main._client_name_from_title(
            "IEC Rocky Mountain (main) | Organic SEO Workbook"
        ) == "IEC Rocky Mountain"

    def test_strips_underscore_bracket_main_and_xlsx_suffix(self):
        assert agent_main._client_name_from_title(
            "Sonoran Spine  _ [Main] Organic SEO Workbook.xlsx"
        ) == "Sonoran Spine"

    def test_plain_name_passes_through_unchanged(self):
        assert agent_main._client_name_from_title("Acme Corp") == "Acme Corp"

    def test_empty_string_returns_empty_string(self):
        assert agent_main._client_name_from_title("") == ""


class TestResolveSheetClientName:
    async def test_prefers_client_details_over_title(self, monkeypatch):
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_client_details",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return(
                {"client": "Sonoran Spine", "website": "https://www.sonoranspine.com/"}
            ),
        )
        result = await agent_main._resolve_sheet_client_name(
            "sheet-id-123", "Sonoran Spine (main) | Organic SEO Workbook"
        )
        assert result == "Sonoran Spine"

    async def test_falls_back_to_title_heuristic_when_no_client_details(self, monkeypatch):
        monkeypatch.setattr(
            agent_main.check_orchestrator, "get_workbook_client_details",
            lambda spreadsheet_id, mcp_url, auth_headers: _async_return({}),
        )
        result = await agent_main._resolve_sheet_client_name(
            "sheet-id-123", "IEC Rocky Mountain (main) | Organic SEO Workbook"
        )
        assert result == "IEC Rocky Mountain"

    async def test_falls_back_to_title_heuristic_on_fetch_failure(self, monkeypatch):
        async def boom(spreadsheet_id, mcp_url, auth_headers):
            raise RuntimeError("no Client Details tab")

        monkeypatch.setattr(agent_main.check_orchestrator, "get_workbook_client_details", boom)
        result = await agent_main._resolve_sheet_client_name(
            "sheet-id-123", "IEC Rocky Mountain (main) | Organic SEO Workbook"
        )
        assert result == "IEC Rocky Mountain"


class TestResolveUploadClientName:
    def test_prefers_client_details_over_filename(self, monkeypatch):
        monkeypatch.setattr(
            agent_main.wu, "parse_client_details_tab",
            lambda file_bytes: {"client": "Sonoran Spine", "website": "https://www.sonoranspine.com/"},
        )
        result = agent_main._resolve_upload_client_name(b"fake-bytes", "workbook.xlsx")
        assert result == "Sonoran Spine"

    def test_falls_back_to_filename_heuristic_when_no_client_details(self, monkeypatch):
        monkeypatch.setattr(agent_main.wu, "parse_client_details_tab", lambda file_bytes: {})
        result = agent_main._resolve_upload_client_name(
            b"fake-bytes", "IEC Rocky Mountain (main) | Organic SEO Workbook.xlsx"
        )
        assert result == "IEC Rocky Mountain"

    def test_falls_back_to_filename_heuristic_on_parse_failure(self, monkeypatch):
        def boom(file_bytes):
            raise ValueError("not a valid xlsx")

        monkeypatch.setattr(agent_main.wu, "parse_client_details_tab", boom)
        result = agent_main._resolve_upload_client_name(
            b"fake-bytes", "IEC Rocky Mountain (main) | Organic SEO Workbook.xlsx"
        )
        assert result == "IEC Rocky Mountain"


class TestRunWorkbookBatch:
    async def test_happy_path_runs_checks_then_submits_report_directly(self, monkeypatch):
        rows = [{"url": "https://example.com/a", "opt_note": "Title Tag"}]
        url_results = [{"url": "https://example.com/a", "verdict": "PASS", "checks": [],
                         "manual_checklist": [], "key_issues": "", "recommended_fixes": ""}]

        async def fake_run_batch(rows_arg, mcp_url, auth_headers, brand_guide=None):
            assert rows_arg == rows
            return url_results, ["Voice & Tone — should be: Friendly"]

        monkeypatch.setattr(agent_main.check_orchestrator, "run_batch", fake_run_batch)

        submit_calls = []

        async def fake_submit_report(client, month_year, results, mcp_url, auth_headers, brand_guide_notes=None):
            submit_calls.append((client, month_year, results, brand_guide_notes))
            return "https://storage.googleapis.com/signed-url"

        monkeypatch.setattr(agent_main.check_orchestrator, "submit_report", fake_submit_report)

        # No LLM step should be reached at all for this path.
        monkeypatch.setattr(
            agent_main, "_run_and_reply",
            lambda *a: (_ for _ in ()).throw(AssertionError("_run_and_reply should not be called")),
        )

        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_workbook_batch(
            "spaces/AAA", "spaces/AAA", "sess1", "August 2025", rows, object(),
            client_name="Acme Corp",
        )

        assert submit_calls == [
            ("Acme Corp", "August 2025", url_results, ["Voice & Tone — should be: Friendly"]),
        ]
        assert "https://storage.googleapis.com/signed-url" in posted["text"]
        assert "All 1 page(s) passed." in posted["text"]

    async def test_summary_reports_fail_count(self, monkeypatch):
        url_results = [
            {"url": "https://example.com/a", "verdict": "FAIL", "checks": [], "manual_checklist": [],
             "key_issues": "", "recommended_fixes": ""},
            {"url": "https://example.com/b", "verdict": "PASS", "checks": [], "manual_checklist": [],
             "key_issues": "", "recommended_fixes": ""},
        ]
        monkeypatch.setattr(
            agent_main.check_orchestrator, "run_batch",
            lambda *a, **kw: _async_return((url_results, [])),
        )
        monkeypatch.setattr(
            agent_main.check_orchestrator, "submit_report",
            lambda *a, **kw: _async_return("https://storage.googleapis.com/signed-url"),
        )
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_workbook_batch(
            "spaces/AAA", "spaces/AAA", "sess1", "August 2025",
            [{"url": "https://example.com/a"}, {"url": "https://example.com/b"}], object(),
        )

        assert "1 passed, 1 need attention." in posted["text"]

    async def test_orchestrator_failure_posts_error_without_calling_llm(self, monkeypatch):
        async def failing_run_batch(rows_arg, mcp_url, auth_headers, brand_guide=None):
            raise RuntimeError("mcp-server unreachable")

        monkeypatch.setattr(agent_main.check_orchestrator, "run_batch", failing_run_batch)

        run_and_reply_calls = []
        monkeypatch.setattr(
            agent_main, "_run_and_reply",
            lambda *a: run_and_reply_calls.append(a) or _async_return(None),
        )

        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_workbook_batch(
            "spaces/AAA", "spaces/AAA", "sess1", "August 2025", [{"url": "https://example.com/a"}], object(),
        )

        assert run_and_reply_calls == []
        assert "mcp-server unreachable" in posted["text"]

    async def test_report_submission_failure_posts_error(self, monkeypatch):
        url_results = [{"url": "https://example.com/a", "verdict": "PASS", "checks": [],
                         "manual_checklist": [], "key_issues": "", "recommended_fixes": ""}]
        monkeypatch.setattr(
            agent_main.check_orchestrator, "run_batch",
            lambda *a, **kw: _async_return((url_results, [])),
        )

        async def failing_submit_report(client, month_year, results, mcp_url, auth_headers, brand_guide_notes=None):
            raise RuntimeError("signing failed")

        monkeypatch.setattr(agent_main.check_orchestrator, "submit_report", failing_submit_report)

        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_workbook_batch(
            "spaces/AAA", "spaces/AAA", "sess1", "August 2025", [{"url": "https://example.com/a"}], object(),
        )

        assert "signing failed" in posted["text"]


# ---------------------------------------------------------------------------
# _download_chat_attachment
# ---------------------------------------------------------------------------

class TestDownloadChatAttachment:
    @respx.mock
    async def test_requests_media_prefixed_path_with_alt_media(self, monkeypatch):
        # Regression test: a bare {resource_name} (no "media/" prefix) 404s —
        # this exact bug shipped once already.
        monkeypatch.setattr(agent_main.google.auth, "default", lambda scopes: (FakeCreds(), None))
        route = respx.get(
            "https://chat.googleapis.com/v1/media/spaces/AAA/messages/BBB/attachments/CCC"
        ).mock(return_value=httpx.Response(200, content=b"xlsx-bytes"))

        result = await agent_main._download_chat_attachment("spaces/AAA/messages/BBB/attachments/CCC")

        assert result == b"xlsx-bytes"
        assert route.called
        assert route.calls[0].request.url.params["alt"] == "media"
        assert route.calls[0].request.headers["authorization"] == "Bearer tok123"


# ---------------------------------------------------------------------------
# _post_to_chat
# ---------------------------------------------------------------------------

class TestPostToChat:
    @respx.mock
    async def test_sends_bearer_token_and_text_payload(self, monkeypatch):
        monkeypatch.setattr(agent_main.google.auth, "default", lambda scopes: (FakeCreds(), None))
        route = respx.post("https://chat.googleapis.com/v1/spaces/AAA/messages").mock(
            return_value=httpx.Response(200, json={})
        )

        await agent_main._post_to_chat("spaces/AAA", "spaces/AAA/threads/T1", "hello world")

        assert route.called
        req = route.calls[0].request
        assert req.headers["authorization"] == "Bearer tok123"
        import json
        payload = json.loads(req.content)
        assert payload["text"] == "hello world"
        assert payload["thread"]["name"] == "spaces/AAA/threads/T1"
        assert "messageReplyOption" not in payload
        assert req.url.params["messageReplyOption"] == "REPLY_MESSAGE_FALLBACK_TO_NEW_THREAD"

    @respx.mock
    async def test_omits_thread_when_thread_equals_space(self, monkeypatch):
        monkeypatch.setattr(agent_main.google.auth, "default", lambda scopes: (FakeCreds(), None))
        route = respx.post("https://chat.googleapis.com/v1/spaces/AAA/messages").mock(
            return_value=httpx.Response(200, json={})
        )

        await agent_main._post_to_chat("spaces/AAA", "spaces/AAA", "hello")

        import json
        req = route.calls[0].request
        payload = json.loads(req.content)
        assert "thread" not in payload
        assert "messageReplyOption" not in req.url.params

    @respx.mock
    async def test_raises_on_non_2xx_response(self, monkeypatch):
        monkeypatch.setattr(agent_main.google.auth, "default", lambda scopes: (FakeCreds(), None))
        respx.post("https://chat.googleapis.com/v1/spaces/AAA/messages").mock(
            return_value=httpx.Response(403, text="forbidden")
        )

        with pytest.raises(httpx.HTTPStatusError):
            await agent_main._post_to_chat("spaces/AAA", "spaces/AAA", "hello")


# ---------------------------------------------------------------------------
# _run_and_reply
# ---------------------------------------------------------------------------

class TestTruncateReply:
    def test_short_reply_is_unchanged(self):
        assert agent_main._truncate_reply("short") == "short"

    def test_no_url_falls_back_to_blind_slice(self):
        long_text = "x" * 5000
        result = agent_main._truncate_reply(long_text)
        assert len(result) <= 4000
        assert result.endswith("_(message truncated — see report link for full details)_")

    def test_preserves_full_report_url_when_truncating(self):
        # The real bug this guards against: a blind reply[:3900] slice landing
        # inside the (very long, query-string-heavy) signed URL corrupts it —
        # GCS then rejects it with SignatureDoesNotMatch.
        url = "https://storage.googleapis.com/bucket/report.html?" + "X-Goog-Signature=" + ("a" * 512)
        reply = ("Here is a very long narrative. " * 200) + f"\n\nReport: {url}"
        assert len(reply) > 4000

        result = agent_main._truncate_reply(reply)

        assert url in result
        assert len(result) <= 4000 + len(url)  # url itself is never cut, even if that alone exceeds 4000
        assert "message truncated" in result

    def test_url_alone_exceeding_limit_is_still_returned_whole(self):
        url = "https://storage.googleapis.com/bucket/report.html?" + ("a" * 5000)
        result = agent_main._truncate_reply(f"Report: {url}")
        assert result == url

    def test_reply_exactly_at_limit_is_unchanged(self):
        reply = "x" * 4000
        assert agent_main._truncate_reply(reply) == reply


class TestRunAndReply:
    async def test_joins_multiple_parts_with_newline(self, monkeypatch):
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(
            agent_main, "Runner",
            make_fake_runner([FakeEvent("part1"), FakeEvent("part2")]),
        )
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert posted["text"] == "part1\npart2"

    async def test_no_final_response_posts_placeholder(self, monkeypatch):
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([]))
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert posted["text"] == "No response generated."

    async def test_logs_diagnostics_when_final_event_has_no_text(self, monkeypatch, caplog):
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([FakeEvent(None)]))

        async def fake_post(space, thread, text):
            pass

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        with caplog.at_level("WARNING"):
            await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert any("Final event had no text" in r.message for r in caplog.records)

    async def test_retries_after_malformed_function_call_and_succeeds(self, monkeypatch):
        # Real production failure: MALFORMED_FUNCTION_CALL on a large Mode B
        # batch's generate_report call produces no text on the first attempt
        # but a retry (fresh session) succeeds.
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(
            agent_main, "Runner",
            make_fake_runner(events_sequence=[[FakeEvent(None)], [FakeEvent("good reply")]]),
        )
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert posted["text"] == "good reply"

    async def test_gives_up_after_max_attempts(self, monkeypatch, caplog):
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        fake_runner_cls = make_fake_runner([FakeEvent(None)])
        monkeypatch.setattr(agent_main, "Runner", fake_runner_cls)
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        with caplog.at_level("WARNING"):
            await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert posted["text"] == "No response generated."
        no_text_warnings = [r for r in caplog.records if "Final event had no text" in r.message]
        assert len(no_text_warnings) == agent_main._RUN_AND_REPLY_MAX_ATTEMPTS

    async def test_first_attempt_success_does_not_retry(self, monkeypatch):
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(
            agent_main, "Runner",
            make_fake_runner(events_sequence=[[FakeEvent("first try")], [FakeEvent("should not be used")]]),
        )
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert posted["text"] == "first try"

    async def test_truncates_long_replies_to_4000_chars(self, monkeypatch):
        long_text = "x" * 5000
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([FakeEvent(long_text)]))
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert len(posted["text"]) <= 4000
        assert posted["text"].startswith("x" * 3900)
        assert posted["text"].endswith("_(message truncated — see report link for full details)_")

    async def test_long_reply_with_report_url_keeps_url_intact(self, monkeypatch):
        url = "https://storage.googleapis.com/bucket/report.html?X-Goog-Signature=" + ("a" * 512)
        long_reply = ("Detailed narrative text. " * 200) + f"\n\nFull report: {url}"
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([FakeEvent(long_reply)]))
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert url in posted["text"]

    async def test_short_replies_are_not_truncated(self, monkeypatch):
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([FakeEvent("short reply")]))
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert posted["text"] == "short reply"

    async def test_agent_exception_is_reported_as_chat_reply(self, monkeypatch):
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner(raise_exc=RuntimeError("kaboom")))
        posted = {}

        async def fake_post(space, thread, text):
            posted["text"] = text

        monkeypatch.setattr(agent_main, "_post_to_chat", fake_post)

        await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())

        assert posted["text"].startswith("❌ Agent error:")
        assert "kaboom" in posted["text"]

    async def test_post_to_chat_failure_does_not_propagate(self, monkeypatch):
        monkeypatch.setattr(agent_main, "create_agent", lambda: object())
        monkeypatch.setattr(agent_main, "Runner", make_fake_runner([FakeEvent("hello")]))

        async def failing_post(space, thread, text):
            raise RuntimeError("chat is down")

        monkeypatch.setattr(agent_main, "_post_to_chat", failing_post)

        # Should not raise despite the Chat API call failing.
        await agent_main._run_and_reply("spaces/AAA", "spaces/AAA", "sess1", "hi", object())
