"""Tests for agent/workbook_upload.py — parsing an uploaded .xlsx into the
same row shape mcp-server/tools/sheets.py produces from a live Google Sheet.
"""
import datetime
import io

import openpyxl
import pytest

import seo_testing_agent.workbook_upload as wu

HEADERS = [
    "Mo - Yr", "Optimization / URL", "What Is Planned / Has Been Done?",
    "Optimization Focus", "Keyword / Volume", "Target Geo", "Inspect in GSC",
    "Is It Live?", "Old Title Tag", "New Title Tag", "Old Meta Description",
    "New Meta Description", "Old H1", "New H1", "Front End Visual QA", "Redirection?",
]


def _build_xlsx(data_rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "On-Page"
    # Rows 1-3 are title/notes rows in the real template — header is row 4.
    ws.append(["ESSENTIAL DETAILS:"])
    ws.append(["- notes"])
    ws.append(["Strategic Plan"])
    ws.append(HEADERS)
    for row in data_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row(month, url, opt_note="", keyword="", geo="Phoenix, AZ", **overrides):
    base = {
        "Mo - Yr": month, "Optimization / URL": url,
        "What Is Planned / Has Been Done?": opt_note,
        "Optimization Focus": "SEO", "Keyword / Volume": keyword,
        "Target Geo": geo, "Inspect in GSC": "", "Is It Live?": "",
        "Old Title Tag": "", "New Title Tag": "", "Old Meta Description": "",
        "New Meta Description": "", "Old H1": "", "New H1": "",
        "Front End Visual QA": "", "Redirection?": "",
    }
    base.update(overrides)
    return [base[h] for h in HEADERS]


class TestGetCol:
    def test_returns_first_matching_alias(self):
        assert wu._get_col({"Month Year": "August 2025"}, "Mo - Yr", "Month Year") == "August 2025"

    def test_prefers_earlier_alias_when_both_present(self):
        row = {"Mo - Yr": "August 2025", "Month Year": "September 2025"}
        assert wu._get_col(row, "Mo - Yr", "Month Year") == "August 2025"

    def test_skips_empty_values_for_earlier_alias(self):
        row = {"Mo - Yr": "", "Month Year": "August 2025"}
        assert wu._get_col(row, "Mo - Yr", "Month Year") == "August 2025"

    def test_returns_empty_string_when_no_alias_matches(self):
        assert wu._get_col({"Other": "x"}, "Mo - Yr", "Month Year") == ""


class TestParseOnPageRows:
    def test_reads_header_and_data_rows(self):
        xlsx = _build_xlsx([
            _row("August 2025", "https://example.com/a", opt_note="Core Opts: Title Tag"),
            _row("August 2025", "https://example.com/b", opt_note="Schema Markup"),
        ])
        rows = wu.parse_on_page_rows(xlsx)
        assert len(rows) == 2
        assert rows[0]["Optimization / URL"] == "https://example.com/a"
        assert rows[1]["What Is Planned / Has Been Done?"] == "Schema Markup"

    def test_skips_fully_blank_rows(self):
        xlsx = _build_xlsx([
            _row("August 2025", "https://example.com/a"),
            ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ])
        rows = wu.parse_on_page_rows(xlsx)
        assert len(rows) == 1

    def test_missing_sheet_raises(self):
        xlsx = _build_xlsx([_row("August 2025", "https://example.com/a")])
        with pytest.raises(ValueError, match="not found"):
            wu.parse_on_page_rows(xlsx, sheet="Nonexistent")

    def test_handles_datetime_month_cell(self):
        xlsx = _build_xlsx([
            _row(datetime.datetime(2025, 10, 1), "https://example.com/c"),
        ])
        rows = wu.parse_on_page_rows(xlsx)
        assert isinstance(rows[0]["Mo - Yr"], datetime.datetime)


class TestDistinctMonths:
    def test_returns_unique_months_in_order(self):
        xlsx = _build_xlsx([
            _row("August 2025", "https://example.com/a"),
            _row("August 2025", "https://example.com/b"),
            _row("September 2025", "https://example.com/c"),
        ])
        rows = wu.parse_on_page_rows(xlsx)
        assert wu.distinct_months(rows) == ["August 2025", "September 2025"]

    def test_skips_month_year_header_separator_rows(self):
        rows = [{"Mo - Yr": "Month Year"}, {"Mo - Yr": "August 2025"}]
        assert wu.distinct_months(rows) == ["August 2025"]

    def test_formats_datetime_months_nicely(self):
        rows = [{"Mo - Yr": datetime.datetime(2025, 10, 1)}]
        assert wu.distinct_months(rows) == ["October 2025"]

    def test_collapses_different_days_in_same_month(self):
        # Real-world data entry inconsistency: same intended month, different day.
        rows = [
            {"Mo - Yr": datetime.datetime(2026, 5, 1)},
            {"Mo - Yr": datetime.datetime(2026, 5, 2)},
        ]
        assert wu.distinct_months(rows) == ["May 2026"]

    def test_falls_back_to_month_year_column_alias(self):
        # Some client workbooks use "Month Year" instead of "Mo - Yr" —
        # seen in the wild on a real IEC Rocky Mountain workbook.
        rows = [{"Month Year": "August 2025"}, {"Month Year": "September 2025"}]
        assert wu.distinct_months(rows) == ["August 2025", "September 2025"]

    def test_prefers_mo_yr_when_both_columns_present(self):
        rows = [{"Mo - Yr": "August 2025", "Month Year": "September 2025"}]
        assert wu.distinct_months(rows) == ["August 2025"]


class TestMonthRows:
    def test_filters_to_requested_month_and_cleans_fields(self):
        xlsx = _build_xlsx([
            _row("August 2025", "https://example.com/a", opt_note="Core Opts",
                 keyword="back pain (170)", geo="Phoenix, AZ"),
            _row("September 2025", "https://example.com/b"),
        ])
        rows = wu.parse_on_page_rows(xlsx)
        result = wu.month_rows(rows, "August 2025")

        assert len(result) == 1
        r = result[0]
        assert r["url"] == "https://example.com/a"
        assert r["keyword"] == "back pain"
        assert r["geo_city"] == "Phoenix"
        assert r["geo_state"] == "AZ"
        assert r["opt_note"] == "Core Opts"

    def test_skips_non_url_rows(self):
        xlsx = _build_xlsx([
            _row("August 2025", "Organic Strategy Renewal"),
            _row("August 2025", "https://example.com/a"),
        ])
        rows = wu.parse_on_page_rows(xlsx)
        result = wu.month_rows(rows, "August 2025")
        assert len(result) == 1
        assert result[0]["url"] == "https://example.com/a"

    def test_unparseable_month_raises(self):
        with pytest.raises(ValueError, match="Could not parse"):
            wu.month_rows([], "not a month")

    def test_visual_qa_bool_coercion(self):
        xlsx = _build_xlsx([
            _row("August 2025", "https://example.com/a", **{"Front End Visual QA": "Yes"}),
        ])
        rows = wu.parse_on_page_rows(xlsx)
        result = wu.month_rows(rows, "August 2025")
        assert result[0]["visual_qa"] is True

    def test_month_year_column_alias(self):
        rows = [{"Month Year": "August 2025", "Optimization / URL": "https://example.com/a"}]
        result = wu.month_rows(rows, "August 2025")
        assert len(result) == 1
        assert result[0]["url"] == "https://example.com/a"

    def test_old_new_headers_column_alias_for_h1(self):
        rows = [{
            "Mo - Yr": "August 2025", "Optimization / URL": "https://example.com/a",
            "Old Headers": "Old H1 Text", "New Headers": "New H1 Text",
        }]
        result = wu.month_rows(rows, "August 2025")
        assert result[0]["old_h1"] == "Old H1 Text"
        assert result[0]["new_h1"] == "New H1 Text"


class TestFindMentionedMonth:
    def test_finds_verbatim_case_insensitive_match(self):
        months = ["August 2025", "September 2025"]
        assert wu.find_mentioned_month("please review august 2025", months) == "August 2025"

    def test_returns_none_when_no_match(self):
        months = ["August 2025", "September 2025"]
        assert wu.find_mentioned_month("please review June 2026", months) is None


def _build_brand_guide_xlsx(rows: list[list], sheet_name: str = "Brand Guide") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "On-Page"
    ws.append(HEADERS)
    bg = wb.create_sheet(sheet_name)
    for row in rows:
        bg.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseBrandGuideTab:
    def test_reads_and_joins_brand_guide_tab(self):
        xlsx = _build_brand_guide_xlsx([
            ["Branding", "Acme Corp"],
            ["CTA", "https://example.com"],
        ])
        result = wu.parse_brand_guide_tab(xlsx)
        assert result == "Branding\tAcme Corp\nCTA\thttps://example.com"

    def test_returns_empty_string_when_tab_missing(self):
        xlsx = _build_xlsx([_row("August 2025", "https://example.com/a")])
        assert wu.parse_brand_guide_tab(xlsx) == ""

    def test_trims_trailing_empty_cells(self):
        xlsx = _build_brand_guide_xlsx([["Branding", "Acme Corp", None, None]])
        assert wu.parse_brand_guide_tab(xlsx) == "Branding\tAcme Corp"

    def test_skips_fully_empty_rows(self):
        xlsx = _build_brand_guide_xlsx([
            ["Branding", "Acme Corp"],
            [None, None],
            ["CTA", "https://example.com"],
        ])
        assert wu.parse_brand_guide_tab(xlsx) == "Branding\tAcme Corp\nCTA\thttps://example.com"


class TestParseClientDetailsTab:
    def test_reads_business_name_and_website(self):
        xlsx = _build_brand_guide_xlsx(
            [["Client Business Name", "Sonoran Spine"], ["Website URL", "https://www.sonoranspine.com/"]],
            sheet_name="Client Details",
        )
        assert wu.parse_client_details_tab(xlsx) == {
            "client": "Sonoran Spine", "website": "https://www.sonoranspine.com/",
        }

    def test_returns_blank_structure_when_tab_missing(self):
        xlsx = _build_xlsx([_row("August 2025", "https://example.com/a")])
        assert wu.parse_client_details_tab(xlsx) == {"client": "", "website": ""}

    def test_ignores_unrecognized_labels(self):
        xlsx = _build_brand_guide_xlsx(
            [["Package Level", "Standard"]], sheet_name="Client Details",
        )
        assert wu.parse_client_details_tab(xlsx) == {"client": "", "website": ""}

    def test_accepts_label_aliases(self):
        xlsx = _build_brand_guide_xlsx(
            [["Client Name", "Acme Corp"]], sheet_name="Client Details",
        )
        assert wu.parse_client_details_tab(xlsx)["client"] == "Acme Corp"
